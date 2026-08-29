#!/usr/bin/env python
"""
compare.py — たたき台v0: 人手完成形xlsx（評価管理表）とAI抽出結果を突合する。

使い方:
  1. python extract.py   を先に実行し、out/fm_import.csv を作る
  2. python compare.py    を実行する

出力（すべて out/ 配下・.gitignore対象）:
  out/mapping_report.md  … fields.yaml の各フィールドと評価管理表ヘッダーの自動対応付け結果
  out/accuracy_report.md … フィールド別・案件別の一致率
                            （一致/不一致/未抽出/対応列なし の件数と率のみ）

★実データの値はどこにも書かない。一致/不一致/未抽出/対応列なし、の判定のみを扱う★

既知の限界（README.md にも記載）:
  - ヘッダー行の検出はヒューリスティック（先頭N行のうち「ラベルらしいセルが最多の行」）。
    結合セルは左上セル以外を空扱いするopenpyxl標準動作のまま。
  - フィールド⇔ヘッダーの対応付けは文字列類似度ベースの単純な処理。
    意味は同じでも表記が大きく異なる列は対応付けに失敗し得る。
  - 一致判定は「AI候補値が、対応付けた列の値のいずれかと一致するか」のみを見ており、
    行単位の対応（どの行と対応するか）までは踏み込まない。
"""
from __future__ import annotations

import csv
import difflib
import sys
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_DIR))

from lib.common import (  # noqa: E402
    ConfigError,
    is_answer_key,
    list_case_files,
    load_config,
    load_fields,
    normalize_value,
    sanitize_error,
)

MATCH_THRESHOLD = 0.5


def is_label_like(v) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip()
    return bool(s) and len(s) <= 30


def find_header_row(ws, scan_rows: int) -> tuple[int, int]:
    # read_only モードでは行末が EmptyCell（.column を持たない簡易セル）で
    # 埋められることがあるため、.value 以外の属性には依存しない。
    max_row = ws.max_row or 1
    best_row_idx = 1
    best_score = -1
    for rno, row in enumerate(ws.iter_rows(min_row=1, max_row=min(scan_rows, max_row)), start=1):
        score = sum(1 for c in row if is_label_like(getattr(c, "value", None)))
        if score > best_score:
            best_score = score
            best_row_idx = rno
    return best_row_idx, best_score


def load_answer_key_columns(path: Path, scan_rows: int):
    """1xlsxの全シートから {source_file, sheet, header, values[]} のリストを作る。"""
    import openpyxl

    columns = []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        return columns, sanitize_error(e)

    try:
        for ws in wb.worksheets:
            if not ws.max_row or ws.max_row < 1:
                continue
            header_row_idx, header_score = find_header_row(ws, scan_rows)
            if header_score < 1:
                continue

            # 列番号は c.column ではなく行内での出現位置（1始まり）で数える。
            # read_only モードの EmptyCell は .column を持たないことがあるため。
            headers: dict[int, str] = {}
            for row in ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx):
                for col_idx, c in enumerate(row, start=1):
                    v = getattr(c, "value", None)
                    if isinstance(v, str) and v.strip():
                        headers[col_idx] = v.strip()
            if not headers:
                continue

            col_values: dict[int, list] = {col: [] for col in headers}
            for row in ws.iter_rows(min_row=header_row_idx + 1):
                for col_idx, c in enumerate(row, start=1):
                    if col_idx in col_values:
                        v = getattr(c, "value", None)
                        if v is not None:
                            col_values[col_idx].append(v)

            for col, header_text in headers.items():
                columns.append(
                    {"source_file": path.name, "sheet": ws.title,
                     "header": header_text, "values": col_values.get(col, [])}
                )
    finally:
        wb.close()

    return columns, None


def match_field_to_columns(field_label: str, columns: list[dict]):
    """類似度最大の列を返す（閾値未満でも返す。閾値クリア判定は呼び出し側で行う）。
    人手レビュー用に「一番近かった候補」を常に mapping_report.md へ残すため。"""
    norm_field = normalize_value(field_label)
    best = None
    best_score = 0.0
    for col in columns:
        norm_header = normalize_value(col["header"])
        if not norm_header:
            continue
        score = difflib.SequenceMatcher(None, norm_field, norm_header).ratio()
        if norm_field and (norm_field in norm_header or norm_header in norm_field):
            score = max(score, 0.8)
        if score > best_score:
            best_score = score
            best = col
    return best, best_score


def load_fm_import(out_dir: Path, csv_name: str, encoding: str):
    path = out_dir / csv_name
    if not path.exists():
        return None
    with open(path, encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        return {row["case_id"]: row for row in reader}


def main():
    try:
        cfg = load_config(PROJECT_DIR)
    except ConfigError as e:
        print(f"[設定エラー] {e}")
        sys.exit(1)

    fields, _doc_kinds = load_fields(PROJECT_DIR)
    out_dir = PROJECT_DIR / "out"

    fm_data = load_fm_import(out_dir, cfg["fm_import_csv_name"], cfg["fm_import_csv_encoding"])
    if fm_data is None:
        print("[エラー] out/fm_import.csv が見つかりません。先に extract.py を実行してください。")
        sys.exit(1)

    scan_rows = cfg["answer_key_header_scan_rows"]
    answer_key_pattern = cfg["answer_key_filename_pattern"]

    mapping_lines = [
        "# mapping_report.md — フィールド ⇔ 評価管理表ヘッダー 自動対応付け結果",
        "",
        "自動対応付けは文字列類似度ベースの単純ヒューリスティックです。"
        "意味的に対応する列でも表記が大きく異なる場合は対応付けに失敗します。"
        "本表は人手確認前提の一次情報として扱ってください。",
    ]

    field_tally = {f.id: {"一致": 0, "不一致": 0, "未抽出": 0, "対応列なし": 0} for f in fields}
    case_tally = {}
    run_errors = []

    for case_id in sorted(cfg["case_map"].keys()):
        real_folder = cfg["case_map"][case_id]
        try:
            all_files = list_case_files(cfg["cases_root"], real_folder)
        except Exception as e:
            run_errors.append(f"{case_id}: {sanitize_error(e)}")
            continue

        ak_files = [p for p in all_files if is_answer_key(p.name, answer_key_pattern)]

        columns: list[dict] = []
        for p in ak_files:
            cols, err = load_answer_key_columns(p, scan_rows)
            if err:
                run_errors.append(f"{case_id}: {err}")
                continue
            columns.extend(cols)

        mapping_lines.append(f"\n## {case_id}")
        mapping_lines.append(f"- 回答キーファイル数: {len(ak_files)} / 検出ヘッダー列数: {len(columns)}")
        mapping_lines.append("")
        mapping_lines.append("| field_id | label | 対応 | 類似度 | 一番近かった候補ヘッダー(参考) |")
        mapping_lines.append("|---|---|---|---|---|")

        ai_row = fm_data.get(case_id, {})
        case_tally[case_id] = {"一致": 0, "不一致": 0, "未抽出": 0, "対応列なし": 0}

        for f in fields:
            best_col, score = match_field_to_columns(f.label_ja, columns) if columns else (None, 0.0)
            mapped = best_col is not None and score >= MATCH_THRESHOLD
            # 閾値未満でも「一番近かった候補」は参考情報として残す（人手レビュー用）。
            best_header_ref = best_col["header"] if best_col is not None else ""
            mapping_lines.append(
                f"| {f.id} | {f.label_ja} | {'mapped' if mapped else 'unmapped'} | "
                f"{score:.2f} | {best_header_ref} |"
            )

            ai_value = normalize_value(ai_row.get(f.id, ""))

            if not mapped:
                verdict = "対応列なし"
            elif not ai_value:
                verdict = "未抽出"
            else:
                manual_values = {normalize_value(v) for v in best_col["values"]}
                verdict = "一致" if ai_value in manual_values else "不一致"

            field_tally[f.id][verdict] += 1
            case_tally[case_id][verdict] += 1

    acc_lines = [
        "# accuracy_report.md — フィールド別・案件別 一致率（v0）",
        "",
        "実データの値はこのファイルに一切書きません。一致/不一致/未抽出/対応列なしの"
        "件数と率のみを記載します。「一致率」は 一致÷(一致+不一致) "
        "（対応付けができ、かつAI側が何か抽出できたケースの中での一致率）です。",
        "",
        "## フィールド別一致率",
        "",
        "| field_id | label | v0_target | 一致 | 不一致 | 未抽出 | 対応列なし | 一致率 |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for f in fields:
        t = field_tally[f.id]
        denom = t["一致"] + t["不一致"]
        rate = f"{(t['一致'] / denom * 100):.0f}%" if denom else "―"
        acc_lines.append(
            f"| {f.id} | {f.label_ja} | {f.v0_target} | {t['一致']} | {t['不一致']} | "
            f"{t['未抽出']} | {t['対応列なし']} | {rate} |"
        )

    acc_lines.append("")
    acc_lines.append("## 案件別一致率")
    acc_lines.append("")
    acc_lines.append("| case_id | 一致 | 不一致 | 未抽出 | 対応列なし | 一致率 |")
    acc_lines.append("|---|---|---|---|---|---|")
    for case_id in sorted(case_tally.keys()):
        t = case_tally[case_id]
        denom = t["一致"] + t["不一致"]
        rate = f"{(t['一致'] / denom * 100):.0f}%" if denom else "―"
        acc_lines.append(f"| {case_id} | {t['一致']} | {t['不一致']} | {t['未抽出']} | {t['対応列なし']} | {rate} |")

    (out_dir / "mapping_report.md").write_text("\n".join(mapping_lines) + "\n", encoding="utf-8")
    (out_dir / "accuracy_report.md").write_text("\n".join(acc_lines) + "\n", encoding="utf-8")

    total_match = sum(t["一致"] for t in field_tally.values())
    total_mismatch = sum(t["不一致"] for t in field_tally.values())
    total_unextracted = sum(t["未抽出"] for t in field_tally.values())
    total_unmapped = sum(t["対応列なし"] for t in field_tally.values())
    denom = total_match + total_mismatch
    overall_rate = f"{(total_match / denom * 100):.0f}%" if denom else "―"

    print(f"案件数: {len(case_tally)}")
    print(
        f"全体集計: 一致={total_match} 不一致={total_mismatch} "
        f"未抽出={total_unextracted} 対応列なし={total_unmapped} 一致率={overall_rate}"
    )
    if run_errors:
        print(f"処理中のエラー: {len(run_errors)}件 → {run_errors}")
    print(f"出力先: {out_dir / 'mapping_report.md'} / {out_dir / 'accuracy_report.md'}")


if __name__ == "__main__":
    main()
